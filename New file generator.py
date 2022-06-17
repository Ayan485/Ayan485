import openpyxl
from openpyxl import Workbook, load_workbook

Name = str(input("Enter file name\n"))
extension = str(input("Enter extension\n"))


filepath = r"C:/Users/Desktop/"+ Name + "." + extension
wb = openpyxl.Workbook()
Sheet_name = wb.sheetnames
wb.save(filepath)

